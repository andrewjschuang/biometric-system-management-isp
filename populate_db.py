# NOME COMPLETO B | SEXO E | DATA NASC F | TELEFONE G | EMAIL I | CULTO K | MEMBRO L | SIGI Q | NOME EM FOTO S

from PIL import Image
import face_recognition
import numpy as np
import openpyxl
import argparse
import os
import io

from Mongodb import Mongodb
import gridfs


class Workbook:
    rows = range(2, 200)
    columns = 'BEFGIKLQS'

    def __init__(self, fp, sheet):
        self.wb = openpyxl.load_workbook(fp)
        self.ws = self.wb.get_sheet_by_name(sheet)

    def list_of_people(self):
        l = []
        for row in Workbook.rows:
            p = []
            for column in Workbook.columns:
                cell = column + str(row)
                value = self.ws[cell].value
                p.append(value)
            if p[0] is not None and p[-1] is not None:
                l.append(p)
        return l

    def dict_of_people(self, l):
        d = []
        for element in l:
            d.append({
                'nome': element[0],
                'sexo': element[1],
                'data_nascimento': element[2],
                'telefone': element[3],
                'email': element[4],
                'ministerio': element[5],
                'membro_isp': element[6],
                'sigi': element[7],
                'fotos': {
                    'central': element[8].lower() + '-fr.jpg',
                    'direita': element[8].lower() + '-ld.jpg',
                    'esquerda': element[8].lower() + '-le.jpg',
                    'obs': element[9] if len(element) == 10 else None
                },
                'images': {
                    'central': element[8].lower() + '-fr.jpg',
                    'direita': element[8].lower() + '-ld.jpg',
                    'esquerda': element[8].lower() + '-le.jpg',
                }
            })
        return d


class Rotate:
    @staticmethod
    def rotate(image):
        image_exif = image._getexif()
        image_orientation = image_exif[274]
        if image_orientation == 2:
            image = image.transpose(Image.FLIP_LEFT_RIGHT)
        if image_orientation == 3:
            image = image.transpose(Image.ROTATE_180)
        if image_orientation == 4:
            image = image.transpose(Image.FLIP_TOP_BOTTOM)
        if image_orientation == 5:
            image = image.transpose(
                Image.FLIP_LEFT_RIGHT).transpose(Image.ROTATE_90)
        if image_orientation == 6:
            image = image.transpose(Image.ROTATE_270)
        if image_orientation == 7:
            image = image.transpose(
                Image.FLIP_TOP_BOTTOM).transpose(Image.ROTATE_90)
        if image_orientation == 8:
            image = image.transpose(Image.ROTATE_90)
        return image


def rename_lower(path):
    for f in os.listdir(path):
        os.rename(os.path.join(path, f), os.path.join(path, f.lower()))


def populate(d, db, args):
    fs = gridfs.GridFS(db.db)
    for person in d:
        encoding_saved = False
        print(person['nome'])
        obs = person['fotos'].pop('obs')

        for key in person['fotos'].keys():
            if person['fotos'][key] is None:
                continue
            else:
                try:
                    fpath = os.path.join(
                        args.path, person['fotos'][key].lower())

                    foto = Image.open(fpath)
                    foto = Rotate.rotate(foto)
                    encodings = face_recognition.face_encodings(np.array(foto))[
                        0]

                    print('saving %s: %s...' % (key, fpath), end=' ')
                    encoding = {
                        'nome': person['nome'],
                        'foto': encodings.tolist(),
                        'obs': obs
                    }

                    encoding_id = db.insert('encodings', encoding)
                    person['fotos'][key] = encoding_id

                    imgByteArr = io.BytesIO()
                    foto.save(imgByteArr, format='JPEG')
                    image_id = fs.put(imgByteArr.getvalue())
                    person['images'][key] = image_id

                    encoding_saved = True
                    print('done')
                except Exception as e:
                    print(e)
                    continue

        if encoding_saved:
            print('saving person...', end=' ')
            member_id = db.insert('members', person)
            print('done\n')
        else:
            print('no encodings saved. skipping...\n')


def createArgsParser():
    parser = argparse.ArgumentParser()
    parser.add_argument('file', help="xlsx file")
    parser.add_argument('-s', '--sheet', required=True, help='xlsx sheet')
    parser.add_argument('-p', '--path', required=True,
                        help='folder with pictures')
    parser.add_argument('-d', '--database', required=True,
                        help='database name')
    return parser.parse_args()


if __name__ == '__main__':
    args = createArgsParser()
    rename_lower(args.path)

    wb = Workbook(args.file, args.sheet)
    d = wb.dict_of_people(wb.list_of_people())

    db = Mongodb(db=args.database)
    db.delete_all('encodings', True)
    db.delete_all('members', True)

    populate(d, db, args)
