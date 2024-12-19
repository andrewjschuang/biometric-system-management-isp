# NOME COMPLETO B | SEXO E | DATA NASC F | TELEFONE G | EMAIL I | CULTO K | MEMBRO L | SIGI Q | NOME EM FOTO S

import os
import re
import openpyxl
import argparse
import requests


class Workbook:
    rows = range(2, 200)
    columns = 'BEFGIKLQS'

    def __init__(self, fp, sheet):
        self.wb = openpyxl.load_workbook(fp)
        self.ws = self.wb.get_sheet_by_name(sheet)

    def parse_wb(self):
        rows = []
        for row in Workbook.rows:
            person = []
            for column in Workbook.columns:
                cell = column + str(row)
                value = self.ws[cell].value
                person.append(value)
            if person[0] is not None and person[-1] is not None:
                rows.append(person)
        return rows

    def parse_rows(self, rows):
        return [self.parse_row(row) for row in rows]

    def parse_row(self, row):
        data = {}

        data['name'] = row[0]

        data['gender'] = 'FEMALE' if row[1] == 'F' else 'MALE'

        # data['birth_date'] = row[2].year, row[2].month, row[2].day if row[2] else None
        data['birth_date'] = '01/01/1990'

        if row[3] is None:
            data['phone_number'] = None
        elif type(row[3]) == float:
            data['phone_number'] = str(row[3])[:-2]
        elif type(row[3]) == int:
            data['phone_number'] = str(row[3])
        else:
            data['phone_number'] = re.compile('[\W_]+').sub('', row[3])

        data['email'] = row[4]

        data['ministry'] = row[5].split(',')[0].upper()

        data['is_member'] = bool(row[6])

        data['sigi'] = row[7]

        data['photos'] = {
            'FRONT': row[8].lower() + '-fr.jpg',
            'LEFT': row[8].lower() + '-le.jpg',
            'RIGHT': row[8].lower() + '-ld.jpg',
        }

        return data


def post(person, base_path):
    path = person['photos']['FRONT']
    photos = { 'FRONT': open(os.path.join(base_path, path), 'rb') }
    del person['photos']

    print(f'Posting {person['name']}')

    response = requests.post('http://localhost:5003/api/members',
                             data=person,
                             files=photos)

    print(f'Status Code: {response.status_code}')


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('file', help="xlsx file")
    parser.add_argument('-s', '--sheet', required=True, help='xlsx sheet')
    parser.add_argument('-p', '--path', required=True,
                        help='folder with pictures')
    args = parser.parse_args()

    for file in os.listdir(args.path):
        os.rename(os.path.join(args.path, file),
                  os.path.join(args.path, file.lower()))

    wb = Workbook(args.file, args.sheet)
    rows = wb.parse_wb()
    people = wb.parse_rows(rows)

    for person in people:
        try:
            post(person, args.path)
        except Exception as e:
            print(f'Error creating {person['name']}: {e}')


if __name__ == '__main__':
    main()
